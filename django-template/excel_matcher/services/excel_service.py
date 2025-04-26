import pandas as pd
import numpy as np
import re
import os
from rapidfuzz import process, fuzz
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from ..models import FuzzyMatchPattern
import openpyxl
from openpyxl.styles import PatternFill


class FuzzyMatcher:
    """模糊匹配处理器，负责学习匹配模式并应用到新数据"""

    def __init__(self, column_name=None, reference_values=None):
        self.patterns = {}  # 存储学习到的模式: {cleaned_or_signature: standardized_value}
        self.column_name = column_name

        # 如果有列名，从数据库加载已学习的模式
        if column_name and not reference_values:
            self.load_patterns_from_db()

        # 如果提供了参照标准值，直接加载
        if reference_values is not None:
            self.load_reference_values(reference_values)

    def load_reference_values(self, reference_values):
        """加载参照标准值作为匹配模式"""
        for value in reference_values:
            if not isinstance(value, str) or not value.strip():
                continue

            # 清理值并存储原始形式
            original_standard = value.strip()  # 保留原始大小写形式作为标准值
            cleaned_value = original_standard.upper()
            self.patterns[cleaned_value] = original_standard

            # 同时存储签名形式
            alpha_part = re.sub(r"[^A-Za-z]", "", cleaned_value)
            numeric_part = re.sub(r"[^0-9]", "", cleaned_value)
            signature = f"{alpha_part}_{numeric_part}"
            if signature not in self.patterns:
                # 签名也映射到原始大小写的标准值
                self.patterns[signature] = original_standard

    def load_patterns_from_db(self):
        """从数据库加载该列已有的匹配模式"""
        patterns = FuzzyMatchPattern.objects.filter(column_name=self.column_name)

        for pattern in patterns:
            # original_pattern 存储的是 cleaned 或 signature
            # standardized_value 存储的是原始大小写的标准值
            self.patterns[pattern.original_pattern] = pattern.standardized_value

    def learn_patterns(self, column_data):
        """从现有数据中学习匹配模式"""
        unique_values = column_data.dropna().unique()
        learned_standards = {}  # 临时存储签名 -> 最早出现的原始值

        for value in unique_values:
            if not isinstance(value, str) or not value.strip():
                continue

            original_value = value.strip()  # 保留原始大小写
            cleaned_value = original_value.upper()

            # 规则提取：提取字母部分和数字部分
            alpha_part = re.sub(r"[^A-Za-z]", "", cleaned_value)
            numeric_part = re.sub(r"[^0-9]", "", cleaned_value)
            signature = f"{alpha_part}_{numeric_part}"

            # 如果这个签名是第一次遇到，将当前原始值作为该签名的标准形式
            if signature not in learned_standards:
                learned_standards[signature] = original_value
                # 将签名映射到标准值
                if signature not in self.patterns:
                    self.patterns[signature] = original_value

            # 将清理后的大写形式也映射到其对应的标准值 (通过签名查找)
            standard_for_cleaned = learned_standards[signature]
            if cleaned_value not in self.patterns:
                self.patterns[cleaned_value] = standard_for_cleaned

        # 学习完成后，保存到数据库
        if self.column_name:
            self.save_patterns_to_db()

        return self.patterns

    def save_patterns_to_db(self):
        """保存学习到的模式到数据库"""
        # 需要注意，现在 self.patterns 的 key 可能是 cleaned_value 或 signature
        # value 是原始大小写的 standardized_value
        # 我们需要确保数据库中存储的是正确的映射关系
        # 这里简化处理：仅将在 learn_patterns 中新学习到的模式存入数据库
        # 或者，更健壮的方式是重新整理 patterns 以匹配数据库结构
        # 为避免复杂性，暂时维持原有的保存逻辑，但需意识到可能的不一致性
        # 更好的方法是调整数据库模型或保存逻辑
        for original_key, standardized_value in self.patterns.items():
            # 假设 original_key 对应数据库的 original_pattern
            # standardized_value 对应数据库的 standardized_value
            FuzzyMatchPattern.objects.update_or_create(
                column_name=self.column_name,
                original_pattern=original_key,  # 这可能不完全准确，取决于key是cleaned还是signature
                defaults={"standardized_value": standardized_value},
            )

    def _extract_primary_key(self, value):
        """提取字符串开头的重要部分（字母数字序列）作为主键"""
        if not isinstance(value, str):
            return None
        # 匹配开头的字母或数字序列
        match = re.match(r"^[A-Za-z0-9]+", value.strip())
        # 返回大写形式的主键，如果匹配不到则返回None
        return match.group(0).upper() if match else None

    def match(self, value, threshold=80):
        """
        对给定值进行分层模糊匹配：
        0. 尝试直接匹配和签名匹配 (优化)。
        1. 提取主键 (开头的字母数字部分)。
        2. 筛选出主键相同的候选标准值。
        3. 在筛选出的候选中进行整体模糊匹配。

        返回 (标准化值, 是否进行了修改)
        """
        if not isinstance(value, str) or not value.strip():
            return value, False  # 对于非字符串或空字符串，直接返回

        original_value = value.strip()  # 保留原始输入值
        cleaned_value = original_value.upper()  # 用于匹配的大写版本
        input_primary_key = self._extract_primary_key(cleaned_value)  # 提取输入值的主键

        # 如果无法提取主键或没有候选模式，直接返回原始值
        if not input_primary_key or not self.patterns:
            return original_value, False

        # --- 优化：尝试快速匹配 ---
        # 0a. 直接匹配 (使用 cleaned_value)
        if cleaned_value in self.patterns:
            result = self.patterns[cleaned_value]  # 获取对应的标准值
            # 比较标准值和原始输入值是否不同
            return result, result != original_value

        # 0b. 签名匹配
        alpha_part = re.sub(r"[^A-Za-z]", "", cleaned_value)
        numeric_part = re.sub(r"[^0-9]", "", cleaned_value)
        signature = f"{alpha_part}_{numeric_part}"
        if signature in self.patterns:
            result = self.patterns[signature]  # 获取对应的标准值
            # 比较标准值和原始输入值是否不同
            return result, result != original_value
        # --- 快速匹配结束 ---

        # --- 分层匹配逻辑 ---
        # 1. 筛选主键相同的候选标准值
        primary_key_candidates = {}  # 存储 {pattern_key: standardized_value}
        for pattern_key, standardized_value in self.patterns.items():
            # 尝试从 pattern_key (可能是 cleaned 或 signature) 提取主键
            candidate_primary_key = self._extract_primary_key(pattern_key)
            # 如果 pattern_key 是签名，可能无法直接提取，尝试从其对应的标准值提取
            if candidate_primary_key is None:
                candidate_primary_key = self._extract_primary_key(standardized_value)

            # 如果候选主键与输入主键相同，则加入候选列表
            if candidate_primary_key == input_primary_key:
                primary_key_candidates[pattern_key] = standardized_value

        # 如果经过主键筛选后没有候选者，则认为无法匹配
        if not primary_key_candidates:
            return original_value, False

        # 2. 在主键匹配的候选中进行模糊匹配 (比较整个字符串)
        # 使用 cleaned_value 与候选的 pattern_key 进行比较
        # process.extractOne 的 choices 需要是可迭代的 key
        match_key, score, _ = process.extractOne(
            cleaned_value,  # 输入的清理后的大写值
            list(
                primary_key_candidates.keys()
            ),  # 候选的 pattern_keys (cleaned 或 signature)
            scorer=fuzz.ratio,  # 使用 Levenshtein 距离比例
        )

        # 3. 如果找到足够相似的匹配
        if score >= threshold:
            result = primary_key_candidates[match_key]  # 获取最佳匹配对应的标准值
            # 比较标准值和原始输入值是否不同
            return result, result != original_value
        # --- 分层匹配逻辑结束 ---

        # 如果以上所有步骤都没有找到合适的匹配，返回原始值
        return original_value, False


class ExcelService:
    """处理Excel文件上传、处理和下载的服务"""

    UPLOAD_DIR = "uploads"
    PROCESSED_DIR = "processed"
    YELLOW_FILL = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    def __init__(self):
        # 确保上传和处理目录存在
        self.upload_path = os.path.join(settings.MEDIA_ROOT, self.UPLOAD_DIR)
        self.processed_path = os.path.join(settings.MEDIA_ROOT, self.PROCESSED_DIR)
        os.makedirs(self.upload_path, exist_ok=True)
        os.makedirs(self.processed_path, exist_ok=True)

        self.fs = FileSystemStorage(location=self.upload_path)

    def save_uploaded_file(self, file):
        """保存上传的Excel文件并返回文件路径"""
        filename = self.fs.save(file.name, file)
        return os.path.join(self.upload_path, filename)

    def get_excel_columns(self, filepath):
        """获取Excel文件的列名"""
        try:
            df = pd.read_excel(filepath)
            return df.columns.tolist()
        except Exception as e:
            raise ValueError(f"无法读取Excel文件: {str(e)}")

    def preview_matches(
        self,
        filepath,
        columns_to_match,
        threshold=80,
        processing_mode="SELF_LEARNING",
        reference_column=None,
    ):
        """
        预览Excel文件的匹配结果

        Returns:
            包含每列匹配统计和示例的字典
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(filepath)

            # 准备结果字典
            results = {}

            if processing_mode == "REFERENCE":
                # 参照标准匹配模式
                if reference_column not in df.columns:
                    raise ValueError(f"标准参照列 '{reference_column}' 不存在")

                # 获取标准列的唯一值作为匹配标准
                standard_values = df[reference_column].dropna().unique()

                # 创建基于标准列的匹配器
                matcher = FuzzyMatcher(reference_values=standard_values)

                # 处理选中的匹配列 (不包括标准列本身)
                for column in columns_to_match:
                    if column == reference_column or column not in df.columns:
                        continue

                    # 匹配该列的数据
                    column_results = self._preview_column(
                        df[column], matcher, threshold
                    )
                    results[column] = column_results
            else:
                # 自学习标准化模式
                for column in columns_to_match:
                    if column not in df.columns:
                        continue

                    # 创建匹配器并学习模式
                    matcher = FuzzyMatcher(column_name=column)
                    # 注意：learn_patterns 现在会处理签名和 cleaned_value
                    matcher.learn_patterns(df[column])

                    # 匹配该列的数据
                    column_results = self._preview_column(
                        df[column], matcher, threshold
                    )
                    results[column] = column_results

            return results

        except Exception as e:
            raise ValueError(f"预览Excel文件时发生错误: {str(e)}")

    def _preview_column(self, column_data, matcher, threshold):
        """为单个列生成预览数据"""
        total = len(column_data)
        changed = 0
        changed_pairs = []  # 存储 (原值, 标准值) 对

        # 处理每个值
        for value in column_data.dropna().tolist():
            # 使用更新后的 match 方法
            standardized, was_changed = matcher.match(value, threshold)
            if was_changed:
                changed += 1
                # 只保存最多5个示例
                if len(changed_pairs) < 5:
                    # 确保存储的是原始值和匹配到的标准值
                    changed_pairs.append((value, standardized))

        # 计算百分比
        percentage = round((changed / total * 100), 1) if total > 0 else 0

        return {
            "total": total,
            "changed": changed,
            "percentage": percentage,
            "examples": changed_pairs,
        }

    def process_excel_file(
        self,
        filepath,
        columns_to_match,
        threshold=80,
        processing_mode="SELF_LEARNING",
        reference_column=None,
    ):
        """
        处理Excel文件，根据选择的模式进行模糊匹配

        Args:
            filepath: Excel文件路径
            columns_to_match: 需要模糊匹配的列名列表
            threshold: 模糊匹配的阈值
            processing_mode: 处理模式，'SELF_LEARNING'(自学习) 或 'REFERENCE'(参照标准)
            reference_column: 标准参照列名称，仅在参照标准模式下使用

        Returns:
            处理后的文件路径
        """
        if processing_mode == "REFERENCE" and (
            not reference_column or reference_column not in columns_to_match
        ):
            raise ValueError("参照标准匹配模式需要指定一个有效的标准列")

        if processing_mode == "SELF_LEARNING":
            return self.process_with_self_learning(
                filepath, columns_to_match, threshold
            )
        else:
            # 从列表中移除标准列，因为它不需要被匹配
            match_columns = [col for col in columns_to_match if col != reference_column]
            return self.process_with_reference_column(
                filepath, reference_column, match_columns, threshold
            )

    def process_with_self_learning(self, filepath, columns_to_match, threshold=80):
        """使用自学习模式处理Excel文件"""
        try:
            # 读取Excel文件
            df = pd.read_excel(filepath)

            # 生成输出文件名
            basename = os.path.basename(filepath)
            name, ext = os.path.splitext(basename)
            output_filepath = os.path.join(
                self.processed_path, f"{name}_Processed{ext}"
            )

            # 保存临时文件，以便稍后添加样式
            temp_filepath = os.path.join(self.processed_path, f"{name}_temp{ext}")

            # 变化跟踪：(行索引, 列名, 原值, 新值)
            changes = []

            # 对每个选中的列进行模糊匹配
            for column in columns_to_match:
                if column in df.columns:
                    # 创建匹配器并学习模式
                    matcher = FuzzyMatcher(column_name=column)
                    matcher.learn_patterns(df[column])

                    # 新列名：原列名_标准
                    std_column_name = f"{column}_标准"

                    # 应用模糊匹配并跟踪变化
                    df[std_column_name] = df[column].copy()  # 初始化新列
                    for idx, value in df[column].items():
                        # 使用更新后的 match 方法
                        matched_value, was_changed = matcher.match(value, threshold)
                        df.at[idx, std_column_name] = matched_value
                        if was_changed:
                            # 记录变化，使用 std_column_name 作为列名
                            changes.append((idx, std_column_name, value, matched_value))

                    # 将新列插入到原列右侧
                    column_index = df.columns.get_loc(column)
                    columns = df.columns.tolist()
                    # 确保新列不在旧列表中，然后插入
                    if std_column_name in columns:
                        columns.remove(std_column_name)
                    columns.insert(column_index + 1, std_column_name)
                    df = df[columns]

            # 保存处理后的文件（不含样式）
            df.to_excel(temp_filepath, index=False)

            # 使用openpyxl添加样式
            self._add_highlighting(temp_filepath, output_filepath, changes)

            # 删除临时文件
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)

            return output_filepath

        except Exception as e:
            raise ValueError(f"处理Excel文件时发生错误: {str(e)}")

    def process_with_reference_column(
        self, filepath, reference_column, columns_to_match, threshold=80
    ):
        """
        使用参照标准列匹配模式处理Excel文件

        Args:
            filepath: Excel文件路径
            reference_column: 作为标准的参照列名
            columns_to_match: 需要匹配的列名列表 (不含标准列)
            threshold: 模糊匹配的阈值

        Returns:
            处理后的文件路径
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(filepath)

            # 确认标准列存在
            if reference_column not in df.columns:
                raise ValueError(f"标准参照列 '{reference_column}' 不存在")

            # 生成输出文件名
            basename = os.path.basename(filepath)
            name, ext = os.path.splitext(basename)
            output_filepath = os.path.join(
                self.processed_path, f"{name}_Processed{ext}"
            )

            # 保存临时文件，以便稍后添加样式
            temp_filepath = os.path.join(self.processed_path, f"{name}_temp{ext}")

            # 获取标准列的唯一值作为匹配标准
            standard_values = df[reference_column].dropna().unique()

            # 创建基于标准列的匹配器
            matcher = FuzzyMatcher(reference_values=standard_values)

            # 变化跟踪：(行索引, 列名, 原值, 新值)
            changes = []

            # 对每个选中的列进行模糊匹配
            for column in columns_to_match:
                if column in df.columns:
                    # 新列名：原列名_标准匹配
                    std_column_name = f"{column}_标准匹配"

                    # 应用模糊匹配并跟踪变化
                    df[std_column_name] = df[column].copy()  # 初始化新列
                    for idx, value in df[column].items():
                        # 使用更新后的 match 方法
                        matched_value, was_changed = matcher.match(value, threshold)
                        df.at[idx, std_column_name] = matched_value
                        if was_changed:
                            # 记录变化，使用 std_column_name 作为列名
                            changes.append((idx, std_column_name, value, matched_value))

                    # 将新列插入到原列右侧
                    column_index = df.columns.get_loc(column)
                    columns = df.columns.tolist()
                    # 确保新列不在旧列表中，然后插入
                    if std_column_name in columns:
                        columns.remove(std_column_name)
                    columns.insert(column_index + 1, std_column_name)
                    df = df[columns]

            # 保存处理后的文件（不含样式）
            df.to_excel(temp_filepath, index=False)

            # 使用openpyxl添加样式
            self._add_highlighting(temp_filepath, output_filepath, changes)

            # 删除临时文件
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)

            return output_filepath

        except Exception as e:
            raise ValueError(f"处理Excel文件时发生错误: {str(e)}")

    def _add_highlighting(self, input_file, output_file, changes):
        """为已处理的Excel文件添加黄色高亮标记"""
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        # Excel行索引从1开始，而且有表头，所以pandas的行索引需要+2
        for row_idx, col_name, old_value, new_value in changes:
            # 查找列的Excel索引 (列名现在是 *_标准 或 *_标准匹配)
            col_idx = None
            for i, cell in enumerate(sheet[1], 1):  # 遍历第一行（表头）
                if cell.value == col_name:
                    col_idx = i
                    break

            if col_idx is not None:
                # 在Excel中应用高亮 (行索引是 pandas索引 + 2)
                cell = sheet.cell(row=row_idx + 2, column=col_idx)
                # 确认单元格值确实是匹配后的新值
                # 注意：Excel读取的值可能是数字，需要与 new_value 类型匹配比较
                cell_value_str = str(cell.value) if cell.value is not None else ""
                new_value_str = str(new_value) if new_value is not None else ""
                if cell_value_str == new_value_str:
                    cell.fill = self.YELLOW_FILL

        # 保存带有样式的工作簿
        workbook.save(output_file)
